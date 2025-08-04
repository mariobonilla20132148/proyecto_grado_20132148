# services.py
from datetime import datetime, timezone, timedelta
from sqlalchemy import select, func
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Importa los modelos desde tu archivo principal de la app.
from app import Empleado, RegistroAsistencia

def get_dashboard_summary_data(db_session):
    """
    Calcula y devuelve los datos de resumen para el dashboard administrativo.
    """
    total_employees = db_session.query(func.count(Empleado.id_empleado)).scalar()
    active_employees = db_session.query(func.count(Empleado.id_empleado)).filter(Empleado.activo == True).scalar()

    # Subconsulta para encontrar el ID del último registro de cada empleado
    latest_record_subquery = select(
        func.max(RegistroAsistencia.id_registro)
    ).group_by(RegistroAsistencia.id_empleado).alias("latest_ids")

    # Contar cuántos de esos últimos registros son de tipo 'entrada'
    clocked_in_count = db_session.query(func.count(RegistroAsistencia.id_registro)).filter(
        RegistroAsistencia.id_registro.in_(latest_record_subquery),
        RegistroAsistencia.tipo == 'entrada'
    ).scalar()
    
    # Obtener los 5 fichajes más recientes con el nombre del empleado
    recent_records_stmt = select(RegistroAsistencia, Empleado.nombre, Empleado.apellido)\
        .join(Empleado, RegistroAsistencia.id_empleado == Empleado.id_empleado)\
        .order_by(RegistroAsistencia.timestamp.desc())\
        .limit(5)
    
    recent_records_results = db_session.execute(recent_records_stmt).all()
    
    recent_records_list = []
    for record, nombre, apellido in recent_records_results:
        rec_dict = record.to_dict()
        rec_dict['nombre_completo'] = f"{nombre} {apellido}"
        recent_records_list.append(rec_dict)

    return {
        "total_employees": total_employees,
        "active_employees": active_employees,
        "clocked_in_count": clocked_in_count,
        "recent_records": recent_records_list
    }


def generate_attendance_report_excel(db_session, employee_dni, year, month, user_tz):
    """
    Genera un reporte de asistencia en Excel para un empleado y mes específicos.
    """
    empleado = db_session.get(Empleado, employee_dni)
    if not empleado:
        raise ValueError("Empleado no encontrado.")
    
    spanish_months = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    spanish_weekdays = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    
    try:
        start_of_month_local = datetime(year, month, 1, tzinfo=user_tz)
        if month == 12:
            end_of_month_local = datetime(year, month, 31, 23, 59, 59, 999999, tzinfo=user_tz)
        else:
            end_of_month_local = datetime(year, month + 1, 1, tzinfo=user_tz) - timedelta(microseconds=1)
        
        start_of_month_utc = start_of_month_local.astimezone(timezone.utc)
        end_of_month_utc = end_of_month_local.astimezone(timezone.utc)
    except ValueError:
        raise ValueError("Fecha de mes/año inválida.")
    
    records_stmt = select(RegistroAsistencia).filter_by(id_empleado=employee_dni).filter(RegistroAsistencia.timestamp.between(start_of_month_utc, end_of_month_utc)).order_by(RegistroAsistencia.timestamp.asc())
    records = db_session.scalars(records_stmt).all()
    
    daily_durations_ms = {}
    employee_active_shifts = {}

    for record in records:
        ts_from_db = record.timestamp
        if ts_from_db.tzinfo is None:
            ts_utc = ts_from_db.replace(tzinfo=timezone.utc)
        else:
            ts_utc = ts_from_db.astimezone(timezone.utc)
        
        local_dt = ts_utc.astimezone(user_tz)

        if record.tipo == 'entrada':
            employee_active_shifts[record.id_empleado] = {'type': 'entrada', 'timestamp': local_dt}
        elif record.tipo == 'salida':
            entry_data = employee_active_shifts.get(record.id_empleado)
            if entry_data and entry_data['type'] == 'entrada':
                entry_timestamp = entry_data['timestamp']
                exit_timestamp = local_dt
                entry_day_key = entry_timestamp.strftime('%Y-%m-%d')
                exit_day_key = exit_timestamp.strftime('%Y-%m-%d')
                if entry_day_key == exit_day_key:
                    duration_ms = (exit_timestamp - entry_timestamp).total_seconds() * 1000
                    daily_durations_ms[entry_day_key] = daily_durations_ms.get(entry_day_key, 0) + duration_ms
                else:
                    boundary_midnight = datetime(exit_timestamp.year, exit_timestamp.month, exit_timestamp.day, 0, 0, 0, tzinfo=user_tz)
                    duration_on_entry_day = boundary_midnight - entry_timestamp
                    daily_durations_ms[entry_day_key] = daily_durations_ms.get(entry_day_key, 0) + (duration_on_entry_day.total_seconds() * 1000)
                    duration_on_exit_day = exit_timestamp - boundary_midnight
                    daily_durations_ms[exit_day_key] = daily_durations_ms.get(exit_day_key, 0) + (duration_on_exit_day.total_seconds() * 1000)
                employee_active_shifts.pop(record.id_empleado)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Reporte {empleado.nombre} {month}-{year}"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    default_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = f"Reporte de Asistencia para {empleado.nombre} {empleado.apellido} (DNI: {empleado.id_empleado})"
    title_cell.font = Font(bold=True, size=16, color="2C3E50")
    title_cell.alignment = default_align
    ws.row_dimensions[1].height = 30
    ws.merge_cells('A2:D2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Mes: {spanish_months[month]} {year}"
    subtitle_cell.font = Font(bold=True, size=12, color="34495E")
    subtitle_cell.alignment = default_align
    ws.row_dimensions[2].height = 20
    ws.append([])
    summary_headers = ["Fecha", "Día de la Semana", "Horas Trabajadas"]
    ws.append(summary_headers)
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = default_align
        cell.border = thin_border
    sorted_daily_summaries = sorted(daily_durations_ms.items())
    for day_str, ms in sorted_daily_summaries:
        day_date = datetime.strptime(day_str, '%Y-%m-%d').date()
        total_hours = int(ms // 3600000)
        total_minutes = int((ms % 3600000) // 60000)
        formatted_duration = f"{total_hours}h {total_minutes}m"
        ws.append([day_date.strftime('%Y-%m-%d'), spanish_weekdays[day_date.weekday()], formatted_duration])
        for cell in ws[ws.max_row]:
            cell.alignment = default_align
            cell.border = thin_border
    total_monthly_ms = sum(daily_durations_ms.values())
    total_monthly_hours = int(total_monthly_ms // 3600000)
    total_monthly_minutes = int((total_monthly_ms % 3600000) // 60000)
    formatted_total_monthly_duration = f"{total_monthly_hours}h {total_monthly_minutes}m"
    total_row_index = ws.max_row + 2
    total_label_cell = ws.cell(row=total_row_index, column=2)
    total_label_cell.value = 'Total Horas del Mes:'
    total_label_cell.font = Font(bold=True, color="34495E")
    total_label_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_label_cell.border = thin_border
    total_value_cell = ws.cell(row=total_row_index, column=3)
    total_value_cell.value = formatted_total_monthly_duration
    total_value_cell.font = Font(bold=True, color="1890FF", size=11)
    total_value_cell.alignment = default_align
    total_value_cell.border = thin_border
    records_title_row_index = total_row_index + 2
    ws.merge_cells(start_row=records_title_row_index, start_column=1, end_row=records_title_row_index, end_column=4)
    records_title_cell = ws.cell(row=records_title_row_index, column=1)
    records_title_cell.value = "Detalle de Registros (Entradas y Salidas)"
    records_title_cell.font = Font(bold=True, size=14, color="34495E")
    records_title_cell.alignment = default_align
    ws.row_dimensions[records_title_row_index].height = 25
    record_headers = ["Fecha", "Hora", "Tipo", "Observaciones"]
    ws.append(record_headers)
    for col_idx, header in enumerate(record_headers, 1):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = default_align
        cell.border = thin_border
    for record in records:
        ts_from_db = record.timestamp
        if ts_from_db.tzinfo is None:
            ts_utc = ts_from_db.replace(tzinfo=timezone.utc)
        else:
            ts_utc = ts_from_db.astimezone(timezone.utc)
        local_timestamp = ts_utc.astimezone(user_tz)
        ws.append([local_timestamp.strftime('%Y-%m-%d'), local_timestamp.strftime('%H:%M:%S'), record.tipo.capitalize(), record.observaciones or ''])
        for col_idx, cell in enumerate(ws[ws.max_row], 1):
            cell.alignment = default_align
            cell.border = thin_border
            if col_idx == 3:
                cell.font = Font(color="28A745" if record.tipo == 'entrada' else "007BFF", bold=True)
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    filename = f"Reporte_Asistencia_{empleado.nombre}_{empleado.apellido}_{year}-{month:02d}.xlsx"
    return file_stream, filename
